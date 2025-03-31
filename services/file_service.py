import os
import shutil
from unidecode import unidecode
from openpyxl import Workbook, load_workbook
from config.settings import settings


def load_or_create_workbook(excel_path: str):

    header = [
        "Vaga",
        "Nome",
        "E-mail",
        "Telefone",
        "Pretensão",
        "Resultado",
        "Motivo",
        "Convite",
        "Mensagem",
    ]

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        wb.save(excel_path)
    return wb, ws


def move_files(
    file_name: str, result_data: dict, approved_path: str, rejected_path: str
):
    """
    Moves and renames PDF files based on evaluation result and contact name.

    Args:
        file_name: Current name of the PDF file
        result_data: Dictionary containing evaluation results and contact info
        approved_path: Path to the approved folder
        rejected_path: Path to the rejected folder
    """
    try:
        source_path = os.path.join(settings.RESUME_PATH, file_name)

        if not os.path.exists(source_path):
            raise FileNotFoundError(f"Source file not found: {source_path}")

        # Get contact name and clean it for filename
        contact_name = result_data["contact"]["name"]
        clean_name = unidecode(contact_name).lower()  # Remove accents and special chars
        clean_name = "".join(
            c if c.isalnum() else "_" for c in clean_name
        )  # Replace spaces/special chars
        clean_name = clean_name.strip("_")  # Remove leading/trailing underscores

        # Create new filename
        base_name, ext = os.path.splitext(file_name)
        new_filename = f"{clean_name}{ext}"

        # Determine destination path based on result
        if result_data.get("result", "").lower() == "aprovado":
            dest_folder = approved_path
        else:
            dest_folder = rejected_path

        # Ensure destination directory exists
        os.makedirs(dest_folder, exist_ok=True)
        dest_path = os.path.join(dest_folder, new_filename)

        # Handle duplicate filenames
        counter = 1
        while os.path.exists(dest_path):
            new_filename = f"{clean_name}_{counter}{ext}"
            dest_path = os.path.join(dest_folder, new_filename)
            counter += 1

        # Move and rename the file
        shutil.move(source_path, dest_path)

    except Exception as e:
        print(f"Error moving/renaming file {file_name}: {e}")
        raise
